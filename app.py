"""
BuMP Monthly Reconciliation Tool — Streamlit UI
================================================

Run locally with:
    streamlit run app.py

Upload Dan's overview workbook and the monthly BuMP report, optionally tweak
the advanced settings, click "Run reconciliation", then download both the
updated Excel workbook and the plain-English markdown summary.

The core logic lives in `reconciliation_engine.py` and is intentionally
UI-agnostic so it can be reused from a CLI, a scheduled job, or an API.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from reconciliation_engine import (
    detect_bump_sheet,
    detect_overview_sheet,
    find_column_by_headers,
    find_difference_column,
    find_latest_bump_column,
    generate_markdown_summary,
    update_overview_workbook,
)

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="BuMP Monthly Reconciliation Tool",
    page_icon="📊",
    layout="wide",
)

st.title("📊 BuMP Monthly Reconciliation Tool")
st.caption(
    "Reconciles Dan's overview workbook against the monthly BuMP finance "
    "report. Updates the *Latest BuMP* and *Difference Since Jan 1st* columns, "
    "preserves the workbook's structure and formatting, and produces a "
    "plain-English markdown summary."
)

# ---------------------------------------------------------------------------
# File uploaders
# ---------------------------------------------------------------------------

col_a, col_b = st.columns(2)
with col_a:
    overview_file = st.file_uploader(
        "1. Upload Dan's overview workbook (.xlsx / .xlsm)",
        type=["xlsx", "xlsm"],
        key="overview_file",
        help="The workbook Dan maintains, with the 'Overview' sheet.",
    )

with col_b:
    bump_file = st.file_uploader(
        "2. Upload this month's BuMP report (.xlsx / .xlsm)",
        type=["xlsx", "xlsm"],
        key="bump_file",
        help="The monthly BuMP export from the WPP finance/commercial team.",
    )

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def _load_workbook_bytes(file_bytes: bytes, keep_vba: bool):
    return load_workbook(io.BytesIO(file_bytes), keep_vba=keep_vba)


def _load_uploaded(uploaded, keep_vba: bool):
    """Read an UploadedFile into an openpyxl Workbook (in memory)."""
    file_bytes = uploaded.getvalue()
    return _load_workbook_bytes(file_bytes, keep_vba=keep_vba)


def _preview_sheet_headers(workbook, sheet_name, n_cols=10):
    sheet = workbook[sheet_name]
    return [
        (get_column_letter(c), sheet.cell(row=1, column=c).value)
        for c in range(1, min(sheet.max_column + 1, n_cols + 1))
    ]


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------

if not overview_file or not bump_file:
    st.info("Upload both files above to continue.")
    st.stop()

try:
    overview_wb = _load_uploaded(
        overview_file, keep_vba=overview_file.name.lower().endswith(".xlsm")
    )
    bump_wb = _load_uploaded(
        bump_file, keep_vba=bump_file.name.lower().endswith(".xlsm")
    )
except Exception as e:
    st.error(f"Could not open one of the files: {e}")
    st.stop()

# Auto-detect sheets and columns for defaults
default_overview_sheet = detect_overview_sheet(overview_wb)
default_bump_sheet = detect_bump_sheet(bump_wb)

# Try to pre-fill output column letters
try:
    _ovr = overview_wb[default_overview_sheet]
    _auto_latest = find_latest_bump_column(_ovr, header_row_top=1)
    _auto_diff = find_difference_column(_ovr, header_row_top=1)
    default_latest_letter = get_column_letter(_auto_latest) if _auto_latest else "H"
    default_diff_letter = get_column_letter(_auto_diff) if _auto_diff else "I"
except Exception:
    default_latest_letter, default_diff_letter = "H", "I"

# Advanced settings
with st.expander("Advanced settings (auto-detected — only change if needed)"):
    c1, c2 = st.columns(2)
    with c1:
        overview_sheet_choice = st.selectbox(
            "Overview sheet",
            options=overview_wb.sheetnames,
            index=overview_wb.sheetnames.index(default_overview_sheet),
        )
        baseline_col_name = st.text_input(
            "Baseline column header (on row 3)",
            value="Sum of Total BAT Budget",
            help="The header text of the post-transition baseline column.",
        )
        latest_bump_letter = st.text_input(
            "Latest BuMP output column (letter)",
            value=default_latest_letter,
            help="Column letter where the summed Latest BuMP value should be written.",
        ).strip().upper()
    with c2:
        bump_sheet_choice = st.selectbox(
            "BuMP sheet",
            options=bump_wb.sheetnames,
            index=bump_wb.sheetnames.index(default_bump_sheet),
        )
        cost_col_name = st.text_input(
            "BuMP cost column",
            value="project_fee_gbp",
            help="The column on the BuMP sheet to sum per scope (always a GBP column).",
        )
        difference_letter = st.text_input(
            "Difference output column (letter)",
            value=default_diff_letter,
            help="Column letter for Latest BuMP minus baseline.",
        ).strip().upper()

    c3, c4 = st.columns(2)
    with c3:
        version_filter = st.text_input(
            "BuMP `version` filter (blank = no filter)", value="CURRENT"
        ).strip() or None
    with c4:
        tracker_filter = st.text_input(
            "BuMP `tracker` filter (blank = no filter)", value="CURRENT"
        ).strip() or None

    clear_unmatched = st.checkbox(
        "Clear unmatched output cells (leave OFF unless you really want to wipe values for scopes missing from this month's BuMP)",
        value=False,
    )

# Show a small preview so the user can sanity-check the detection
with st.expander("Sheet detection preview"):
    st.write(f"**Overview sheet:** `{overview_sheet_choice}`")
    st.write("First-row headers:")
    st.write(_preview_sheet_headers(overview_wb, overview_sheet_choice))
    st.write(f"**BuMP sheet:** `{bump_sheet_choice}`")
    st.write("First-row headers:")
    st.write(_preview_sheet_headers(bump_wb, bump_sheet_choice))

st.divider()

# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if not st.button("▶️ Run reconciliation", type="primary", use_container_width=True):
    st.stop()

try:
    result = update_overview_workbook(
        overview_wb=overview_wb,
        bump_wb=bump_wb,
        overview_sheet_name=overview_sheet_choice,
        bump_sheet_name=bump_sheet_choice,
        cost_column=cost_col_name,
        baseline_column_name=baseline_col_name,
        latest_bump_col_letter=latest_bump_letter or None,
        difference_col_letter=difference_letter or None,
        clear_unmatched=clear_unmatched,
        version_filter=version_filter,
        tracker_filter=tracker_filter,
    )
except ValueError as e:
    st.error(f"❌ Reconciliation could not start: {e}")
    st.stop()
except Exception as e:
    st.exception(e)
    st.stop()

# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------

found = [u for u in result.scope_updates if u.latest_bump is not None]
not_found = [u for u in result.scope_updates if u.latest_bump is None]

st.success("Reconciliation complete.")

c1, c2, c3, c4 = st.columns(4)
c1.metric("Scopes in overview", len(result.scope_updates))
c2.metric("Found in BuMP", len(found))
c3.metric("Not in BuMP", len(not_found))
c4.metric("New BuMP scopes", len(result.unmatched_bump_scopes))

total_latest = sum((u.latest_bump for u in found), Decimal("0"))
total_baseline = sum((u.baseline for u in found if u.baseline is not None), Decimal("0"))
net_movement = total_latest - total_baseline

c5, c6 = st.columns(2)
c5.metric("Total Latest BuMP value", f"£{total_latest:,.2f}")
c6.metric(
    "Net movement vs baseline",
    f"£{net_movement:,.2f}",
    delta=f"{(net_movement / total_baseline * 100):.1f}%" if total_baseline else None,
)

# Scope-level table
st.subheader("Per-scope detail")
import pandas as pd

rows = []
for u in result.scope_updates:
    rows.append({
        "Row": u.row,
        "Scope ID": u.scope_id,
        "Brand": u.brand or "",
        "Market": u.bat_market or "",
        "WPP Country": u.country or "",
        "Baseline": float(u.baseline) if u.baseline is not None else None,
        "Latest BuMP": float(u.latest_bump) if u.latest_bump is not None else None,
        "Difference": float(u.difference) if u.difference is not None else None,
        "BuMP rows": u.bump_row_count,
        "Notes": u.note,
    })
df = pd.DataFrame(rows)
st.dataframe(
    df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Baseline": st.column_config.NumberColumn(format="£%.2f"),
        "Latest BuMP": st.column_config.NumberColumn(format="£%.2f"),
        "Difference": st.column_config.NumberColumn(format="£%.2f"),
    },
)

if result.unmatched_bump_scopes:
    st.subheader("New / unmatched BuMP scopes")
    new_rows = [{
        "Scope ID": u.scope_id,
        "Brand": u.brand or "",
        "Market": u.bat_market or "",
        "WPP Country": u.country or "",
        "Project": u.project or "",
        "Latest BuMP Value": float(u.latest_bump),
        "BuMP rows": u.row_count,
    } for u in result.unmatched_bump_scopes]
    st.dataframe(
        pd.DataFrame(new_rows),
        use_container_width=True,
        hide_index=True,
        column_config={"Latest BuMP Value": st.column_config.NumberColumn(format="£%.2f")},
    )

if result.warnings:
    with st.expander(f"⚠️ Data quality warnings ({len(result.warnings)})"):
        for w in result.warnings:
            st.markdown(f"- {w}")

# Markdown preview
markdown = generate_markdown_summary(result)
with st.expander("📝 Markdown summary preview"):
    st.markdown(markdown)

# ---------------------------------------------------------------------------
# Downloads
# ---------------------------------------------------------------------------

st.subheader("Downloads")

# Save workbook to memory
buf = io.BytesIO()
overview_wb.save(buf)
buf.seek(0)

today_str = date.today().isoformat()
xlsx_name = f"Dan_Reconciliation_Updated_{today_str}.xlsx"
md_name = f"Dan_Reconciliation_Summary_{today_str}.md"

dl1, dl2 = st.columns(2)
with dl1:
    st.download_button(
        label="⬇️ Download updated Excel workbook",
        data=buf.getvalue(),
        file_name=xlsx_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with dl2:
    st.download_button(
        label="⬇️ Download markdown summary",
        data=markdown.encode("utf-8"),
        file_name=md_name,
        mime="text/markdown",
        use_container_width=True,
    )
