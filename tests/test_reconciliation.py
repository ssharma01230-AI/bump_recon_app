"""
Unit tests for reconciliation_engine.

Run with: python -m pytest tests/ -v
"""
from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

# Allow running tests from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest
from openpyxl import Workbook

from reconciliation_engine import (
    build_bump_scope_summary,
    classify_overview_rows,
    detect_bump_sheet,
    detect_header_row,
    detect_overview_sheet,
    find_column_by_headers,
    find_difference_column,
    find_latest_bump_column,
    generate_markdown_summary,
    normalise_scope_id,
    update_overview_workbook,
)


# ---------------------------------------------------------------------------
# normalise_scope_id
# ---------------------------------------------------------------------------

class TestNormaliseScopeId:
    def test_int(self):
        assert normalise_scope_id(4705) == "4705"

    def test_string(self):
        assert normalise_scope_id("4705") == "4705"

    def test_float_integer_valued(self):
        assert normalise_scope_id(4705.0) == "4705"

    def test_string_with_trailing_dot_zero(self):
        assert normalise_scope_id("4705.0") == "4705"

    def test_whitespace_stripped(self):
        assert normalise_scope_id("  4705  ") == "4705"

    def test_none(self):
        assert normalise_scope_id(None) == ""

    def test_empty_string(self):
        assert normalise_scope_id("") == ""

    def test_leading_zeros_preserved(self):
        assert normalise_scope_id("04705") == "04705"

    def test_int_and_string_equivalent(self):
        assert normalise_scope_id(4705) == normalise_scope_id("4705")


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def _make_overview_workbook():
    wb = Workbook()
    wb.active.title = "Overview"
    sh = wb["Overview"]
    sh["A1"] = "BAT MARKET"
    sh["D1"] = "SCOPE ID"
    sh["F1"] = "Post Transition"
    sh["H1"] = "LATEST BUMP\n 2026-04-14"
    sh["I1"] = "DIFFERENCE SINCE JAN 1st"
    sh["A3"] = "bat_market"
    sh["B3"] = "country"
    sh["C3"] = "Brand"
    sh["D3"] = "scope_id"
    sh["F3"] = "Sum of Total BAT Budget"
    sh["H3"] = "Sum of Total BAT \nBudget 14th April 2026"
    sh["I3"] = '"+ / -"'
    return wb


def _make_bump_workbook():
    wb = Workbook()
    wb.active.title = "BUMP_2026-04-14"
    sh = wb["BUMP_2026-04-14"]
    headers = [
        "scope_id", "version", "tracker", "bat_market", "country",
        "project_brand", "brand", "project", "project_fee_gbp", "latest_version",
    ]
    for i, h in enumerate(headers, 1):
        sh.cell(row=1, column=i).value = h
    return wb, headers


class TestSheetDetection:
    def test_detect_overview_sheet_by_name(self):
        wb = _make_overview_workbook()
        wb.create_sheet("Sheet2")
        assert detect_overview_sheet(wb) == "Overview"

    def test_detect_bump_sheet_by_scope_id_header(self):
        wb, _ = _make_bump_workbook()
        wb.create_sheet("Notes")
        assert detect_bump_sheet(wb) == "BUMP_2026-04-14"

    def test_detect_header_row_finds_correct_row(self):
        wb = _make_overview_workbook()
        sh = wb["Overview"]
        assert detect_header_row(sh, ["scope_id", "Sum of Total BAT Budget"]) == 3

    def test_detect_header_row_raises_if_missing(self):
        wb = _make_overview_workbook()
        sh = wb["Overview"]
        with pytest.raises(ValueError):
            detect_header_row(sh, ["nonexistent_column"])

    def test_find_latest_bump_column(self):
        wb = _make_overview_workbook()
        sh = wb["Overview"]
        assert find_latest_bump_column(sh, header_row_top=1) == 8

    def test_find_difference_column(self):
        wb = _make_overview_workbook()
        sh = wb["Overview"]
        assert find_difference_column(sh, header_row_top=1) == 9


# ---------------------------------------------------------------------------
# build_bump_scope_summary
# ---------------------------------------------------------------------------

class TestBuildBumpScopeSummary:
    def _populate(self, sh, rows):
        for i, row in enumerate(rows, start=2):
            for j, val in enumerate(row, start=1):
                sh.cell(row=i, column=j).value = val

    def test_basic_sum(self):
        wb, _ = _make_bump_workbook()
        sh = wb["BUMP_2026-04-14"]
        # scope_id, version, tracker, bat_market, country, project_brand, brand, project, fee_gbp, latest
        self._populate(sh, [
            ["4733", "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 100.50, "v1"],
            ["4733", "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 50.25, "v1"],
            ["4807", "CURRENT", "CURRENT", "Belgium", "France", "VUSE", "VUSE", "P2", 999.99, "v1"],
        ])
        summary = build_bump_scope_summary(sh)
        assert summary["4733"]["sum_gbp"] == Decimal("150.75")
        assert summary["4733"]["row_count"] == 2
        assert summary["4807"]["sum_gbp"] == Decimal("999.99")
        assert summary["4807"]["row_count"] == 1

    def test_filters_non_current_versions(self):
        wb, _ = _make_bump_workbook()
        sh = wb["BUMP_2026-04-14"]
        self._populate(sh, [
            ["4733", "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 100.0, "v1"],
            ["4733", "OLD",     "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 999.0, "v1"],
            ["4733", "CURRENT", "PREV",    "Denmark", "Denmark", "VUSE", "VUSE", "P1", 999.0, "v1"],
        ])
        summary = build_bump_scope_summary(sh)
        assert summary["4733"]["sum_gbp"] == Decimal("100.0")
        assert summary["4733"]["row_count"] == 1
        assert summary["4733"]["excluded_rows"] == 2

    def test_scope_ids_normalised(self):
        wb, _ = _make_bump_workbook()
        sh = wb["BUMP_2026-04-14"]
        self._populate(sh, [
            [4733, "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 100.0, "v1"],
            ["4733", "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 50.0, "v1"],
        ])
        summary = build_bump_scope_summary(sh)
        assert "4733" in summary
        assert summary["4733"]["sum_gbp"] == Decimal("150.0")

    def test_missing_cost_column_raises(self):
        wb, _ = _make_bump_workbook()
        sh = wb["BUMP_2026-04-14"]
        with pytest.raises(ValueError, match="missing required column"):
            build_bump_scope_summary(sh, cost_column="nonexistent_col")


# ---------------------------------------------------------------------------
# classify_overview_rows
# ---------------------------------------------------------------------------

def _populate_overview_with_data(wb):
    sh = wb["Overview"]
    # data rows starting row 4
    rows = [
        ("Belgium",  "France", "MULTI BRAND", 4604, 2282.57),
        ("Belgium",  "France", "VUSE", 4807, 15818.48),
        ("Belgium Total", None, None, None, 18101.05),
        (None, None, None, None, None),
        ("Denmark",  "Denmark", "VUSE", 4733, 9236.71),
        ("Denmark Total", None, None, None, 9236.71),
        (None, None, None, None, None),
        ("WESTERN EUROPE", "Romania", "GLO", 4531, 408157.84),
        ("WESTERN EUROPE", "Romania Total", None, None, 408157.84),
        ("WESTERN EUROPE Total", None, None, None, 408157.84),
        ("Grand Total", None, None, None, 435496.07),
    ]
    for i, (a, b, c, d, f) in enumerate(rows, start=4):
        sh.cell(row=i, column=1).value = a
        sh.cell(row=i, column=2).value = b
        sh.cell(row=i, column=3).value = c
        sh.cell(row=i, column=4).value = d
        sh.cell(row=i, column=6).value = f
    return sh


class TestClassifyOverviewRows:
    def test_scope_rows_and_totals_split_correctly(self):
        wb = _make_overview_workbook()
        sh = _populate_overview_with_data(wb)
        scope_rows, totals = classify_overview_rows(sh, header_row=3, col_scope=4)
        assert sorted(scope_rows) == [4, 5, 8, 11]
        labels = [t["label"] for t in totals]
        assert "Belgium Total" in labels
        assert "Denmark Total" in labels
        assert "Romania Total" in labels
        assert "WESTERN EUROPE Total" in labels
        assert "Grand Total" in labels

    def test_market_total_groups_only_its_market(self):
        wb = _make_overview_workbook()
        _populate_overview_with_data(wb)
        sh = wb["Overview"]
        scope_rows, totals = classify_overview_rows(sh, header_row=3, col_scope=4)
        belgium_total = [t for t in totals if t["label"] == "Belgium Total"][0]
        # Belgium rows are rows 4 and 5
        assert sorted(belgium_total["scope_rows"]) == [4, 5]
        denmark_total = [t for t in totals if t["label"] == "Denmark Total"][0]
        assert denmark_total["scope_rows"] == [8]
        grand = [t for t in totals if t["label"] == "Grand Total"][0]
        assert sorted(grand["scope_rows"]) == [4, 5, 8, 11]

    def test_subgroup_total_uses_column_b(self):
        wb = _make_overview_workbook()
        _populate_overview_with_data(wb)
        sh = wb["Overview"]
        _, totals = classify_overview_rows(sh, header_row=3, col_scope=4)
        rom = [t for t in totals if t["label"] == "Romania Total"][0]
        assert rom["scope_rows"] == [11]


# ---------------------------------------------------------------------------
# End-to-end with a fully synthetic pair of workbooks
# ---------------------------------------------------------------------------

class TestEndToEnd:
    def _build_pair(self):
        ovr = _make_overview_workbook()
        sh = ovr["Overview"]
        # Two scopes: 4733 will match, 9999 will not
        sh.cell(row=4, column=1).value = "Denmark"
        sh.cell(row=4, column=2).value = "Denmark"
        sh.cell(row=4, column=3).value = "VUSE"
        sh.cell(row=4, column=4).value = 4733
        sh.cell(row=4, column=6).value = 9236.71

        sh.cell(row=5, column=1).value = "Denmark"
        sh.cell(row=5, column=2).value = "Denmark"
        sh.cell(row=5, column=3).value = "MULTI BRAND"
        sh.cell(row=5, column=4).value = 9999
        sh.cell(row=5, column=6).value = 1000.00
        sh.cell(row=5, column=8).value = 555.55  # existing value
        sh.cell(row=5, column=9).value = -444.45

        sh.cell(row=6, column=1).value = "Denmark Total"
        sh.cell(row=6, column=6).value = 10236.71

        bmp, _ = _make_bump_workbook()
        bsh = bmp["BUMP_2026-04-14"]
        for i, row in enumerate([
            [4733, "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 5000.00, "v1"],
            [4733, "CURRENT", "CURRENT", "Denmark", "Denmark", "VUSE", "VUSE", "P1", 4242.87, "v1"],
            # An unmatched-in-Dan scope
            [5500, "CURRENT", "CURRENT", "Spain", "Spain", "GLO", "GLO", "P2", 1234.56, "v1"],
        ], start=2):
            for j, val in enumerate(row, start=1):
                bsh.cell(row=i, column=j).value = val
        return ovr, bmp

    def test_matched_scope_is_updated(self):
        ovr, bmp = self._build_pair()
        result = update_overview_workbook(ovr, bmp)
        upd = [u for u in result.scope_updates if u.scope_id == "4733"][0]
        assert upd.latest_bump == Decimal("9242.87")
        assert upd.difference == Decimal("6.16")
        # cell actually written
        assert ovr["Overview"].cell(row=4, column=8).value == 9242.87
        assert ovr["Overview"].cell(row=4, column=9).value == 6.16

    def test_unmatched_scope_left_unchanged_by_default(self):
        ovr, bmp = self._build_pair()
        update_overview_workbook(ovr, bmp)
        # Row 5 had pre-existing 555.55 and -444.45; should still be there
        assert ovr["Overview"].cell(row=5, column=8).value == 555.55
        assert ovr["Overview"].cell(row=5, column=9).value == -444.45

    def test_clear_unmatched_blanks_cells(self):
        ovr, bmp = self._build_pair()
        update_overview_workbook(ovr, bmp, clear_unmatched=True)
        assert ovr["Overview"].cell(row=5, column=8).value is None
        assert ovr["Overview"].cell(row=5, column=9).value is None

    def test_new_bump_scope_reported(self):
        ovr, bmp = self._build_pair()
        result = update_overview_workbook(ovr, bmp)
        assert len(result.unmatched_bump_scopes) == 1
        new = result.unmatched_bump_scopes[0]
        assert new.scope_id == "5500"
        assert new.latest_bump == Decimal("1234.56")

    def test_column_headers_unchanged(self):
        ovr, bmp = self._build_pair()
        before_row1 = [ovr["Overview"].cell(row=1, column=c).value for c in range(1, 10)]
        before_row3 = [ovr["Overview"].cell(row=3, column=c).value for c in range(1, 10)]
        update_overview_workbook(ovr, bmp)
        after_row1 = [ovr["Overview"].cell(row=1, column=c).value for c in range(1, 10)]
        after_row3 = [ovr["Overview"].cell(row=3, column=c).value for c in range(1, 10)]
        assert before_row1 == after_row1
        assert before_row3 == after_row3

    def test_total_row_label_unchanged(self):
        ovr, bmp = self._build_pair()
        update_overview_workbook(ovr, bmp)
        assert ovr["Overview"].cell(row=6, column=1).value == "Denmark Total"

    def test_markdown_summary_renders(self):
        ovr, bmp = self._build_pair()
        result = update_overview_workbook(ovr, bmp)
        md = generate_markdown_summary(result)
        assert "Monthly BuMP Reconciliation Summary" in md
        assert "4733" in md
        assert "9999" in md  # unmatched
        assert "5500" in md  # new


# ---------------------------------------------------------------------------
# Acceptance tests against the real uploaded files (skipped if absent)
# ---------------------------------------------------------------------------

REAL_OVERVIEW = Path("/home/claude/Dan_s_reconciliation_report.xlsx")
REAL_BUMP = Path("/home/claude/bump_monthyl_report.xlsx")


@pytest.mark.skipif(
    not (REAL_OVERVIEW.exists() and REAL_BUMP.exists()),
    reason="real workbooks not present in working directory",
)
class TestRealWorkbooks:
    def test_known_scope_calculation(self):
        from openpyxl import load_workbook
        ovr = load_workbook(REAL_OVERVIEW)
        bmp = load_workbook(REAL_BUMP)
        result = update_overview_workbook(ovr, bmp)

        # Scope 4733: baseline 9236.71, expected latest 9242.87, diff +6.16
        upd = [u for u in result.scope_updates if u.scope_id == "4733"][0]
        assert upd.latest_bump == Decimal("9242.87")
        assert upd.difference == Decimal("6.16")

    def test_acceptance_overall(self):
        from openpyxl import load_workbook
        ovr = load_workbook(REAL_OVERVIEW)
        bmp = load_workbook(REAL_BUMP)
        result = update_overview_workbook(ovr, bmp)
        # Expect at least one match, at least one new BuMP scope
        assert any(u.latest_bump is not None for u in result.scope_updates)
        assert len(result.unmatched_bump_scopes) > 0
        # No formula errors anywhere — we never write formulas, only values
        # Column labels untouched
        ovr_sheet = ovr["Overview"]
        assert ovr_sheet["A1"].value == "BAT MARKET"
        assert ovr_sheet["D1"].value == "SCOPE ID"
        assert ovr_sheet["A3"].value == "bat_market"
        assert ovr_sheet["D3"].value == "scope_id"
