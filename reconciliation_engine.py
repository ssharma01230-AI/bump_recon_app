"""
BuMP Monthly Reconciliation Engine
==================================

Pure logic for reconciling Dan's overview workbook against the monthly BuMP
finance report. Deliberately decoupled from any UI so it can be reused from a
Streamlit app, a CLI, a scheduled job, or a REST API.

Public surface
--------------
- detect_overview_sheet(workbook) -> str
- detect_bump_sheet(workbook) -> str
- detect_header_row(sheet, required_headers) -> int
- normalise_scope_id(value) -> str
- build_bump_scope_summary(sheet, cost_column, version_filter, tracker_filter) -> dict
- update_overview_workbook(...) -> ReconciliationResult
- generate_markdown_summary(result) -> str
- recalculate_total_rows(sheet, ...) -> list[TotalRowUpdate]

The reconciliation is "non-destructive by default": rows that don't match a BuMP
scope are left as-is, totals are only recalculated when their grouping can be
unambiguously inferred, and original column/row labels are never modified.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ScopeUpdate:
    """One scope-level row in Dan's overview that we attempted to update."""
    row: int
    scope_id: str
    bat_market: str | None
    country: str | None
    brand: str | None
    baseline: Decimal | None
    latest_bump: Decimal | None          # None => not found in BuMP
    difference: Decimal | None
    bump_row_count: int = 0
    note: str = ""
    # Cross-check info from BuMP (for data-quality warnings)
    bump_bat_market: str | None = None
    bump_country: str | None = None
    bump_brand: str | None = None


@dataclass
class UnmatchedBumpScope:
    """A scope present in BuMP but absent from Dan's overview."""
    scope_id: str
    bat_market: str | None
    country: str | None
    brand: str | None
    project: str | None
    latest_bump: Decimal
    row_count: int


@dataclass
class TotalRowUpdate:
    """A total/subtotal row that was recalculated."""
    row: int
    label: str
    new_latest_bump: Decimal | None
    new_difference: Decimal | None
    scope_rows_included: list[int] = field(default_factory=list)
    skipped_reason: str | None = None


@dataclass
class ReconciliationResult:
    """Bundle returned to the caller (UI/CLI/API)."""
    scope_updates: list[ScopeUpdate]
    unmatched_bump_scopes: list[UnmatchedBumpScope]
    total_updates: list[TotalRowUpdate]
    warnings: list[str]
    # Column/row diagnostics
    overview_sheet_name: str
    bump_sheet_name: str
    header_row: int
    cost_column_name: str
    baseline_column_name: str
    latest_bump_column_letter: str
    difference_column_letter: str


# ---------------------------------------------------------------------------
# Detection helpers
# ---------------------------------------------------------------------------

def _norm_header(value) -> str:
    """Normalise a header cell for fuzzy matching."""
    if value is None:
        return ""
    return " ".join(str(value).lower().replace("\n", " ").split())


def detect_overview_sheet(workbook: Workbook) -> str:
    """
    Return the most likely overview sheet name.

    Preference order:
      1. A sheet literally called "Overview" (case-insensitive).
      2. The first sheet whose row 3 contains "scope_id" (Dan's pivot layout).
      3. The first sheet whose first 5 rows contain "scope_id" anywhere.
      4. Fallback: the first sheet.
    """
    for name in workbook.sheetnames:
        if name.strip().lower() == "overview":
            return name

    for name in workbook.sheetnames:
        sheet = workbook[name]
        for row_idx in (3, 1, 2):
            if row_idx > sheet.max_row:
                continue
            for cell in sheet[row_idx]:
                if _norm_header(cell.value) == "scope_id":
                    return name

    for name in workbook.sheetnames:
        sheet = workbook[name]
        for r in range(1, min(6, sheet.max_row + 1)):
            for cell in sheet[r]:
                if _norm_header(cell.value) == "scope_id":
                    return name

    return workbook.sheetnames[0]


def detect_bump_sheet(workbook: Workbook) -> str:
    """
    Find the sheet that looks like the monthly BuMP export.

    Strategy: pick the sheet whose row 1 contains "scope_id" AND has the
    largest number of columns (BuMP exports are wide, ~100 cols).
    """
    candidates: list[tuple[str, int]] = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        if sheet.max_row < 1:
            continue
        header_cells = [_norm_header(c.value) for c in sheet[1]]
        if "scope_id" in header_cells:
            candidates.append((name, sheet.max_column))

    if candidates:
        candidates.sort(key=lambda x: x[1], reverse=True)
        return candidates[0][0]

    # Fall back to name heuristic
    for name in workbook.sheetnames:
        if name.upper().startswith("BUMP"):
            return name

    return workbook.sheetnames[0]


def detect_header_row(sheet: Worksheet, required_headers: Iterable[str]) -> int:
    """
    Find the row (1-indexed) on `sheet` that contains all `required_headers`.

    Headers are matched after normalisation (case-insensitive, whitespace and
    newlines collapsed). Searches the first 10 rows.
    """
    required = {_norm_header(h) for h in required_headers}
    for r in range(1, min(11, sheet.max_row + 1)):
        row_headers = {_norm_header(c.value) for c in sheet[r] if c.value is not None}
        if required.issubset(row_headers):
            return r
    raise ValueError(
        f"Could not locate a header row containing all of: {sorted(required)}"
    )


def find_column_by_headers(
    sheet: Worksheet, header_row: int, candidates: Iterable[str], contains: bool = False
) -> int | None:
    """
    Return the 1-indexed column number of the first cell on `header_row` that
    matches one of `candidates`. If `contains` is True, match by substring.
    """
    candidates_norm = [_norm_header(c) for c in candidates]
    for c in range(1, sheet.max_column + 1):
        cell_val = _norm_header(sheet.cell(row=header_row, column=c).value)
        if not cell_val:
            continue
        if contains:
            if any(cand in cell_val for cand in candidates_norm):
                return c
        else:
            if cell_val in candidates_norm:
                return c
    return None


def find_latest_bump_column(sheet: Worksheet, header_row_top: int = 1) -> int | None:
    """
    Locate the "Latest BuMP" output column on Dan's overview.

    Looks on the visible header row (row 1 in Dan's file) for cells containing
    'latest bump' or 'total bat' + 'budget' + a date.
    """
    for c in range(1, sheet.max_column + 1):
        val = _norm_header(sheet.cell(row=header_row_top, column=c).value)
        if "latest bump" in val:
            return c
        if "total bat" in val and "budget" in val and any(ch.isdigit() for ch in val):
            return c
    return None


def find_difference_column(sheet: Worksheet, header_row_top: int = 1) -> int | None:
    """Locate the difference / +/- column."""
    for c in range(1, sheet.max_column + 1):
        val = _norm_header(sheet.cell(row=header_row_top, column=c).value)
        if "difference since" in val:
            return c
        if "+ / -" in val or "+/-" in val:
            return c
    return None


# ---------------------------------------------------------------------------
# Scope ID normalisation
# ---------------------------------------------------------------------------

def normalise_scope_id(value) -> str:
    """
    Convert any scope-id-like value to a canonical string.

    Treats `4705`, `"4705"`, `4705.0`, and `" 4705 "` as the same scope.
    Preserves leading zeros if present in the original string. Returns ""
    for None / empty / NaN-ish values.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        # Floats from Excel: integer-valued ones become "4705" not "4705.0"
        if value != value:  # NaN
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s.endswith(".0"):
        # Pandas-style int-as-float string
        rest = s[:-2]
        if rest.lstrip("-").isdigit():
            return rest
    return s


# ---------------------------------------------------------------------------
# BuMP aggregation
# ---------------------------------------------------------------------------

def build_bump_scope_summary(
    sheet: Worksheet,
    cost_column: str = "project_fee_gbp",
    version_filter: str | None = "CURRENT",
    tracker_filter: str | None = "CURRENT",
) -> dict[str, dict]:
    """
    Aggregate the BuMP sheet into a per-scope summary.

    Returns: { scope_id: {
        "sum_gbp": Decimal,
        "row_count": int,
        "bat_market": str | None,
        "country": str | None,
        "brand": str | None,
        "project_brand": str | None,
        "project": str | None,
        "excluded_rows": int,        # filtered out by version/tracker
    } }
    """
    headers = {
        _norm_header(sheet.cell(row=1, column=c).value): c
        for c in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=c).value is not None
    }

    required = ["scope_id", _norm_header(cost_column)]
    missing = [r for r in required if r not in headers]
    if missing:
        raise ValueError(
            f"BuMP sheet is missing required column(s): {missing}. "
            f"Found columns: {list(headers.keys())[:20]}..."
        )

    col_scope = headers["scope_id"]
    col_cost = headers[_norm_header(cost_column)]
    col_version = headers.get("version")
    col_tracker = headers.get("tracker")
    col_market = headers.get("bat_market")
    col_country = headers.get("country")
    col_brand = headers.get("brand")
    col_project_brand = headers.get("project_brand")
    col_project = headers.get("project")

    summary: dict[str, dict] = {}

    for r in range(2, sheet.max_row + 1):
        scope = normalise_scope_id(sheet.cell(row=r, column=col_scope).value)
        if not scope:
            continue

        # Apply CURRENT filters if columns exist
        excluded = False
        if version_filter is not None and col_version is not None:
            v = sheet.cell(row=r, column=col_version).value
            if v is None or str(v).strip().upper() != version_filter.upper():
                excluded = True
        if not excluded and tracker_filter is not None and col_tracker is not None:
            t = sheet.cell(row=r, column=col_tracker).value
            if t is None or str(t).strip().upper() != tracker_filter.upper():
                excluded = True

        entry = summary.setdefault(scope, {
            "sum_gbp": Decimal("0"),
            "row_count": 0,
            "bat_market": None,
            "country": None,
            "brand": None,
            "project_brand": None,
            "project": None,
            "excluded_rows": 0,
        })

        if excluded:
            entry["excluded_rows"] += 1
            continue

        cost_val = sheet.cell(row=r, column=col_cost).value
        if cost_val is not None:
            try:
                entry["sum_gbp"] += Decimal(str(cost_val))
            except Exception:
                # non-numeric, skip but flag elsewhere
                pass
        entry["row_count"] += 1

        # Capture first non-empty descriptor seen
        if entry["bat_market"] is None and col_market:
            entry["bat_market"] = sheet.cell(row=r, column=col_market).value
        if entry["country"] is None and col_country:
            entry["country"] = sheet.cell(row=r, column=col_country).value
        if entry["brand"] is None and col_brand:
            entry["brand"] = sheet.cell(row=r, column=col_brand).value
        if entry["project_brand"] is None and col_project_brand:
            entry["project_brand"] = sheet.cell(row=r, column=col_project_brand).value
        if entry["project"] is None and col_project:
            entry["project"] = sheet.cell(row=r, column=col_project).value

    return summary


# ---------------------------------------------------------------------------
# Total-row detection and recalculation
# ---------------------------------------------------------------------------

def _is_total_label(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return s.lower().endswith(" total") or s.lower() == "grand total"


def classify_overview_rows(
    sheet: Worksheet,
    header_row: int,
    col_scope: int,
    col_market: int = 1,
    col_country: int = 2,
) -> tuple[list[int], list[dict]]:
    """
    Walk the overview sheet and split rows into two lists:
      - scope_rows: rows with a real scope_id (the "data" rows)
      - total_rows: rows that look like subtotals/grand totals, with metadata
                    about which scope rows roll up into them.

    The grouping rule for totals:
      - If the total label is in column A (col_market):
            label "X Total" totals scope rows above this point whose col A == "X"
            (stopping at the previous total row).
      - If the total label is in column B (col_country):
            sub-group within column A — totals the immediately-preceding contiguous
            run of scope rows where (col_A, col_B) match the parent group.
      - "Grand Total": totals all preceding scope rows.

    Returns (scope_rows, total_rows_meta).
    """
    scope_rows: list[int] = []
    totals: list[dict] = []

    for r in range(header_row + 1, sheet.max_row + 1):
        scope_val = sheet.cell(row=r, column=col_scope).value
        a_val = sheet.cell(row=r, column=col_market).value
        b_val = sheet.cell(row=r, column=col_country).value

        if normalise_scope_id(scope_val):
            scope_rows.append(r)
            continue

        # Total candidates
        if _is_total_label(a_val) and not normalise_scope_id(scope_val):
            label = str(a_val).strip()
            if label.lower() == "grand total":
                totals.append({"row": r, "label": label, "kind": "grand", "match": None})
            elif label.lower() == "western europe total":
                totals.append({"row": r, "label": label, "kind": "market", "match": "WESTERN EUROPE"})
            else:
                market = label[:-len(" Total")].strip()
                totals.append({"row": r, "label": label, "kind": "market", "match": market})
            continue

        if _is_total_label(b_val) and not normalise_scope_id(scope_val):
            label = str(b_val).strip()
            sub = label[:-len(" Total")].strip()
            parent = str(a_val).strip() if a_val is not None else None
            totals.append({
                "row": r, "label": label, "kind": "subgroup",
                "match": sub, "parent": parent,
            })
            continue

    # Assign which scope rows belong to each total
    for total in totals:
        rows_for_total: list[int] = []
        if total["kind"] == "grand":
            rows_for_total = list(scope_rows)
        elif total["kind"] == "market":
            wanted = total["match"]
            for sr in scope_rows:
                if sr >= total["row"]:
                    break
                a = sheet.cell(row=sr, column=col_market).value
                if a is not None and str(a).strip() == wanted:
                    rows_for_total.append(sr)
        elif total["kind"] == "subgroup":
            wanted_a = total.get("parent")
            wanted_b = total["match"]
            for sr in scope_rows:
                if sr >= total["row"]:
                    break
                a = sheet.cell(row=sr, column=col_market).value
                b = sheet.cell(row=sr, column=col_country).value
                if (
                    a is not None and str(a).strip() == wanted_a
                    and b is not None and str(b).strip() == wanted_b
                ):
                    rows_for_total.append(sr)

        total["scope_rows"] = rows_for_total

    return scope_rows, totals


# ---------------------------------------------------------------------------
# Main reconciliation
# ---------------------------------------------------------------------------

def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        d = Decimal(str(value))
        if d != d:  # NaN
            return None
        return d
    except Exception:
        return None


def _round2(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def update_overview_workbook(
    overview_wb: Workbook,
    bump_wb: Workbook,
    overview_sheet_name: str | None = None,
    bump_sheet_name: str | None = None,
    cost_column: str = "project_fee_gbp",
    baseline_column_name: str = "Sum of Total BAT Budget",
    latest_bump_col_letter: str | None = None,
    difference_col_letter: str | None = None,
    clear_unmatched: bool = False,
    version_filter: str | None = "CURRENT",
    tracker_filter: str | None = "CURRENT",
) -> ReconciliationResult:
    """
    Run reconciliation and mutate `overview_wb` in place.

    Workbook formatting (number formats, fonts, fills, formulas elsewhere) is
    preserved because we only assign values to specific target cells.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    warnings: list[str] = []

    # --- Sheets ---
    if overview_sheet_name is None:
        overview_sheet_name = detect_overview_sheet(overview_wb)
    if bump_sheet_name is None:
        bump_sheet_name = detect_bump_sheet(bump_wb)
    overview = overview_wb[overview_sheet_name]
    bump = bump_wb[bump_sheet_name]

    # --- Header row + columns on overview ---
    header_row = detect_header_row(
        overview, required_headers=["scope_id", baseline_column_name]
    )
    col_scope = find_column_by_headers(overview, header_row, ["scope_id"])
    col_baseline = find_column_by_headers(overview, header_row, [baseline_column_name])
    if col_scope is None:
        raise ValueError("Could not find 'scope_id' column on the overview sheet.")
    if col_baseline is None:
        raise ValueError(
            f"Could not find baseline column '{baseline_column_name}' on the overview sheet."
        )

    # Latest BuMP and difference output columns
    if latest_bump_col_letter:
        col_latest = column_index_from_string(latest_bump_col_letter)
    else:
        col_latest = find_latest_bump_column(overview, header_row_top=1)
        if col_latest is None:
            raise ValueError(
                "Could not auto-detect the Latest BuMP output column. "
                "Please specify it manually via latest_bump_col_letter."
            )

    if difference_col_letter:
        col_diff = column_index_from_string(difference_col_letter)
    else:
        col_diff = find_difference_column(overview, header_row_top=1)
        if col_diff is None:
            raise ValueError(
                "Could not auto-detect the Difference column. "
                "Please specify it manually via difference_col_letter."
            )

    # --- BuMP summary ---
    bump_summary = build_bump_scope_summary(
        bump, cost_column=cost_column,
        version_filter=version_filter, tracker_filter=tracker_filter,
    )

    # --- Walk overview scope rows ---
    scope_rows, total_rows_meta = classify_overview_rows(
        overview, header_row=header_row, col_scope=col_scope
    )

    scope_updates: list[ScopeUpdate] = []
    seen_dan_scopes: dict[str, int] = {}

    for r in scope_rows:
        scope = normalise_scope_id(overview.cell(row=r, column=col_scope).value)
        if scope in seen_dan_scopes:
            warnings.append(
                f"Duplicate Scope ID `{scope}` in Dan's overview at rows "
                f"{seen_dan_scopes[scope]} and {r}."
            )
        else:
            seen_dan_scopes[scope] = r

        bat_market = overview.cell(row=r, column=1).value
        country = overview.cell(row=r, column=2).value
        brand = overview.cell(row=r, column=3).value
        baseline = _to_decimal(overview.cell(row=r, column=col_baseline).value)

        bump_entry = bump_summary.get(scope)

        if bump_entry is None or bump_entry["row_count"] == 0:
            # Not found
            update = ScopeUpdate(
                row=r, scope_id=scope,
                bat_market=str(bat_market) if bat_market else None,
                country=str(country) if country else None,
                brand=str(brand) if brand else None,
                baseline=baseline,
                latest_bump=None,
                difference=None,
                bump_row_count=0,
                note="Scope not found in this month's BuMP report.",
            )
            if clear_unmatched:
                overview.cell(row=r, column=col_latest).value = None
                overview.cell(row=r, column=col_diff).value = None
            scope_updates.append(update)
            continue

        latest_bump = _round2(bump_entry["sum_gbp"])
        difference = None
        if baseline is not None:
            difference = _round2(latest_bump - baseline)

        # Write values (preserves number_format, font, fill etc.)
        overview.cell(row=r, column=col_latest).value = float(latest_bump)
        if difference is not None:
            overview.cell(row=r, column=col_diff).value = float(difference)
        elif clear_unmatched:
            overview.cell(row=r, column=col_diff).value = None

        # Cross-check market/brand
        cross_warn = []
        bm = bump_entry["bat_market"]
        bc = bump_entry["country"]
        bb = bump_entry["brand"] or bump_entry["project_brand"]
        if bm and bat_market and str(bm).strip().lower() != str(bat_market).strip().lower():
            cross_warn.append(f"bat_market differs (overview='{bat_market}', BuMP='{bm}')")
        if bc and country and str(bc).strip().lower() != str(country).strip().lower():
            cross_warn.append(f"country differs (overview='{country}', BuMP='{bc}')")
        if bb and brand and str(bb).strip().lower() != str(brand).strip().lower():
            cross_warn.append(f"brand differs (overview='{brand}', BuMP='{bb}')")
        if cross_warn:
            warnings.append(f"Scope `{scope}` (row {r}): " + "; ".join(cross_warn))

        if bump_entry["excluded_rows"]:
            warnings.append(
                f"Scope `{scope}`: {bump_entry['excluded_rows']} BuMP row(s) "
                f"excluded by version/tracker filter."
            )

        if baseline is None:
            note = "Baseline is missing in overview; latest BuMP written but no difference calculated."
        elif baseline == Decimal("0"):
            note = "Baseline is zero — latest BuMP recorded as full new amount."
        else:
            note = _plain_english_note(latest_bump, baseline, difference)

        scope_updates.append(ScopeUpdate(
            row=r, scope_id=scope,
            bat_market=str(bat_market) if bat_market else None,
            country=str(country) if country else None,
            brand=str(brand) if brand else None,
            baseline=baseline, latest_bump=latest_bump, difference=difference,
            bump_row_count=bump_entry["row_count"],
            note=note,
            bump_bat_market=str(bm) if bm else None,
            bump_country=str(bc) if bc else None,
            bump_brand=str(bb) if bb else None,
        ))

    # --- Unmatched BuMP scopes ---
    dan_scopes = set(seen_dan_scopes.keys())
    unmatched: list[UnmatchedBumpScope] = []
    for scope, entry in bump_summary.items():
        if scope in dan_scopes:
            continue
        if entry["row_count"] == 0:
            continue
        unmatched.append(UnmatchedBumpScope(
            scope_id=scope,
            bat_market=str(entry["bat_market"]) if entry["bat_market"] else None,
            country=str(entry["country"]) if entry["country"] else None,
            brand=str(entry["brand"] or entry["project_brand"] or "") or None,
            project=str(entry["project"]) if entry["project"] else None,
            latest_bump=_round2(entry["sum_gbp"]),
            row_count=entry["row_count"],
        ))
    unmatched.sort(key=lambda u: u.latest_bump, reverse=True)

    # --- Total rows ---
    total_updates = recalculate_total_rows(
        overview, total_rows_meta, scope_updates,
        col_latest=col_latest, col_diff=col_diff,
    )

    return ReconciliationResult(
        scope_updates=scope_updates,
        unmatched_bump_scopes=unmatched,
        total_updates=total_updates,
        warnings=warnings,
        overview_sheet_name=overview_sheet_name,
        bump_sheet_name=bump_sheet_name,
        header_row=header_row,
        cost_column_name=cost_column,
        baseline_column_name=baseline_column_name,
        latest_bump_column_letter=get_column_letter(col_latest),
        difference_column_letter=get_column_letter(col_diff),
    )


def _plain_english_note(latest: Decimal, baseline: Decimal, diff: Decimal | None) -> str:
    if diff is None:
        return "No baseline to compare against."
    abs_diff = abs(diff)
    if baseline > 0:
        pct = abs_diff / baseline * Decimal(100)
    else:
        pct = Decimal("0")

    if abs_diff < Decimal("1.00") or pct < Decimal("0.5"):
        return "Latest BuMP is broadly aligned with baseline."
    if diff > 0:
        return f"Latest BuMP is £{abs_diff:,.2f} higher than baseline."
    return f"Latest BuMP is £{abs_diff:,.2f} lower than baseline."


def recalculate_total_rows(
    overview: Worksheet,
    totals_meta: list[dict],
    scope_updates: list[ScopeUpdate],
    col_latest: int,
    col_diff: int,
) -> list[TotalRowUpdate]:
    """
    Recompute each total row's Latest BuMP and Difference cells.

    Only writes a new value when *every* scope row in the group has been
    successfully updated this month (i.e. matched in BuMP and has a baseline
    so a difference is defined). If any scope in the group is missing, the
    total is left as-is and a skipped_reason is recorded.
    """
    updates_by_row = {u.row: u for u in scope_updates}
    result: list[TotalRowUpdate] = []

    for meta in totals_meta:
        row = meta["row"]
        rows_in_group = meta.get("scope_rows", [])
        if not rows_in_group:
            result.append(TotalRowUpdate(
                row=row, label=meta["label"],
                new_latest_bump=None, new_difference=None,
                skipped_reason="No scope rows could be associated with this total."
            ))
            continue

        # Sum: include every scope row in the group. For rows where Latest BuMP
        # was not refreshed this month (unmatched in BuMP), fall back to the
        # value already in the cell so the total stays consistent. Blank cells
        # are treated as zero, matching Excel's normal SUM() semantics — this
        # is what users expect from a total row.
        sum_latest = Decimal("0")
        sum_diff = Decimal("0")
        for sr in rows_in_group:
            upd = updates_by_row.get(sr)
            if upd and upd.latest_bump is not None:
                sum_latest += upd.latest_bump
            else:
                cell_val = overview.cell(row=sr, column=col_latest).value
                d = _to_decimal(cell_val)
                if d is not None:
                    sum_latest += d
            if upd and upd.difference is not None:
                sum_diff += upd.difference
            else:
                cell_val = overview.cell(row=sr, column=col_diff).value
                d = _to_decimal(cell_val)
                if d is not None:
                    sum_diff += d

        sum_latest = _round2(sum_latest)
        sum_diff = _round2(sum_diff)
        overview.cell(row=row, column=col_latest).value = float(sum_latest)
        overview.cell(row=row, column=col_diff).value = float(sum_diff)

        result.append(TotalRowUpdate(
            row=row, label=meta["label"],
            new_latest_bump=sum_latest, new_difference=sum_diff,
            scope_rows_included=rows_in_group,
        ))

    return result


# ---------------------------------------------------------------------------
# Markdown summary
# ---------------------------------------------------------------------------

def _fmt_money(d: Decimal | None) -> str:
    if d is None:
        return "—"
    sign = "-" if d < 0 else ""
    return f"{sign}£{abs(d):,.2f}"


def generate_markdown_summary(result: ReconciliationResult) -> str:
    """Render the reconciliation result as a markdown report."""
    found = [u for u in result.scope_updates if u.latest_bump is not None]
    not_found = [u for u in result.scope_updates if u.latest_bump is None]

    total_latest = sum((u.latest_bump for u in found), Decimal("0"))
    total_baseline_for_found = sum(
        (u.baseline for u in found if u.baseline is not None), Decimal("0")
    )
    net_movement = total_latest - total_baseline_for_found

    diffs = [u for u in found if u.difference is not None]
    diffs_sorted_asc = sorted(diffs, key=lambda u: u.difference)
    largest_decreases = diffs_sorted_asc[:5]
    largest_increases = list(reversed(diffs_sorted_asc[-5:]))

    lines: list[str] = []
    lines.append("# Monthly BuMP Reconciliation Summary")
    lines.append("")
    lines.append(f"_Overview sheet: **{result.overview_sheet_name}** · "
                 f"BuMP sheet: **{result.bump_sheet_name}** · "
                 f"Cost column: **{result.cost_column_name}** · "
                 f"Baseline column: **{result.baseline_column_name}**_")
    lines.append("")

    # Executive summary
    lines.append("## Executive Summary")
    lines.append("")
    lines.append(f"- **Total scopes in Dan's overview:** {len(result.scope_updates)}")
    lines.append(f"- **Scopes found in this month's BuMP report:** {len(found)}")
    lines.append(f"- **Scopes not found in this month's BuMP report:** {len(not_found)}")
    lines.append(f"- **New BuMP scopes not in Dan's overview:** {len(result.unmatched_bump_scopes)}")
    lines.append(f"- **Total Latest BuMP value updated:** {_fmt_money(total_latest)}")
    lines.append(f"- **Net movement vs baseline (matched scopes only):** {_fmt_money(net_movement)}")
    lines.append("")

    if largest_increases:
        lines.append("**Largest increases vs baseline:**")
        for u in largest_increases:
            if u.difference is None or u.difference <= 0:
                continue
            lines.append(
                f"- Scope `{u.scope_id}` ({u.brand or '—'}, {u.bat_market or '—'}): "
                f"{_fmt_money(u.difference)}"
            )
        lines.append("")

    if largest_decreases:
        lines.append("**Largest decreases vs baseline:**")
        for u in largest_decreases:
            if u.difference is None or u.difference >= 0:
                continue
            lines.append(
                f"- Scope `{u.scope_id}` ({u.brand or '—'}, {u.bat_market or '—'}): "
                f"{_fmt_money(u.difference)}"
            )
        lines.append("")

    # This month's breakdown
    lines.append("## This Month's Breakdown")
    lines.append("")
    lines.append("| Scope ID | Brand | Market | WPP Country | Baseline | Latest BuMP | Difference | Notes |")
    lines.append("|---|---|---|---|---:|---:|---:|---|")
    for u in result.scope_updates:
        lines.append(
            f"| {u.scope_id} | {u.brand or '—'} | {u.bat_market or '—'} | "
            f"{u.country or '—'} | {_fmt_money(u.baseline)} | "
            f"{_fmt_money(u.latest_bump)} | {_fmt_money(u.difference)} | "
            f"{u.note} |"
        )
    lines.append("")

    # Not found
    lines.append("## Not Found in Monthly BuMP Report")
    lines.append("")
    if not_found:
        lines.append("| Scope ID | Brand | Market | WPP Country | Baseline | Notes |")
        lines.append("|---|---|---|---|---:|---|")
        for u in not_found:
            lines.append(
                f"| {u.scope_id} | {u.brand or '—'} | {u.bat_market or '—'} | "
                f"{u.country or '—'} | {_fmt_money(u.baseline)} | "
                f"Scope not in this month's BuMP report; existing values left unchanged. |"
            )
    else:
        lines.append("_All scopes in Dan's overview were found in the BuMP report._")
    lines.append("")

    # New / unmatched BuMP scopes
    lines.append("## New / Unmatched BuMP Scopes")
    lines.append("")
    if result.unmatched_bump_scopes:
        lines.append("| Scope ID | Brand | Market | WPP Country | Project | Latest BuMP Value | Rows | Notes |")
        lines.append("|---|---|---|---|---|---:|---:|---|")
        for u in result.unmatched_bump_scopes:
            lines.append(
                f"| {u.scope_id} | {u.brand or '—'} | {u.bat_market or '—'} | "
                f"{u.country or '—'} | {u.project or '—'} | "
                f"{_fmt_money(u.latest_bump)} | {u.row_count} | "
                f"New scope in BuMP; not present in Dan's overview. |"
            )
    else:
        lines.append("_No new BuMP scopes outside Dan's overview._")
    lines.append("")

    # Total rows
    if result.total_updates:
        lines.append("## Total Rows")
        lines.append("")
        lines.append("| Row | Label | Latest BuMP | Difference | Status |")
        lines.append("|---:|---|---:|---:|---|")
        for t in result.total_updates:
            status = "Updated"
            if t.skipped_reason:
                status = f"Left unchanged — {t.skipped_reason}"
            lines.append(
                f"| {t.row} | {t.label} | {_fmt_money(t.new_latest_bump)} | "
                f"{_fmt_money(t.new_difference)} | {status} |"
            )
        lines.append("")

    # Data quality
    lines.append("## Data Quality Checks")
    lines.append("")
    quality_items: list[str] = list(result.warnings)
    missing_baseline = [u for u in result.scope_updates if u.baseline is None]
    if missing_baseline:
        for u in missing_baseline:
            quality_items.append(f"Scope `{u.scope_id}` (row {u.row}) has no baseline value.")
    skipped_totals = [t for t in result.total_updates if t.skipped_reason]
    for t in skipped_totals:
        quality_items.append(f"Total row {t.row} ('{t.label}') was not recalculated: {t.skipped_reason}")

    if quality_items:
        for w in quality_items:
            lines.append(f"- {w}")
    else:
        lines.append("_No data quality issues detected._")
    lines.append("")

    return "\n".join(lines)
